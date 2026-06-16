#!/usr/bin/env python3
"""Generate the TaloNet procurement BOM as a professional multi-sheet .xlsx.

Reproducible: edit the data tables below and re-run to regenerate
``bom/TaloNet_BOM.xlsx``. Covers the net-interceptor drone ("그물매") and the
ForensIQ-1 forensic appliance, with MAKE / BUY / OUTSOURCE disposition, indicative
USD pricing, vendors, and a formula-driven cost roll-up. Part numbers trace to
docs/01 (drone spec), docs/06 (circuit BOM), docs/09 (appliance), and
gcs/payload_map.py (servo channels).

    python tools/generate_bom.py
"""

from __future__ import annotations

import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

KHAKI = "86855F"
DARK = "23241C"
GREEN = "DDEAD0"   # MAKE
BLUE = "D6E4F0"    # BUY
AMBER = "F6E6C4"   # OUTSOURCE
GREY = "ECECE8"    # TOOL

COLS = ["ID", "Category", "Item / Description", "Mfr / Part No.", "Qty", "Unit",
        "Disp", "Unit USD", "Ext USD", "Vendor", "Ref", "Notes"]
WIDTHS = [11, 20, 40, 26, 6, 7, 11, 11, 12, 18, 9, 40]

thin = Side(style="thin", color="B8B8AE")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
DISP_FILL = {"MAKE": GREEN, "BUY": BLUE, "OUTSOURCE": AMBER, "TOOL": GREY}


# ---------------------------------------------------------------------------
# DATA  (id, category, item, mpn, qty, unit, disp, unit_usd, vendor, ref, notes)
# ---------------------------------------------------------------------------
DRONE_FRAME = [
    ("DR-AF-001", "Frame", "Hex carbon deck plate, 360 mm AF, 3 mm (top/bottom)", "TaloNet GM-DECK (CNC)", 2, "ea", "OUTSOURCE", 130, "Local CNC shop", "cad", "From talonet_frame.scad; CF/G10, waterjet/CNC"),
    ("DR-AF-002", "Frame", "Carbon arm tube 30x26 mm, 420 mm", "Rock West 30x26", 4, "ea", "BUY", 22, "Rock West Composites", "cad", "Roll-wrapped CF; X8 = 4 arms"),
    ("DR-AF-003", "Frame", "Arm clamp + coaxial motor mount (alu 6061)", "TaloNet GM-MNT (CNC)", 4, "ea", "OUTSOURCE", 38, "Local CNC shop", "cad", "Holds 2 coaxial motors/arm"),
    ("DR-AF-004", "Frame", "Standoff/spacer set, alu, deck stack", "M3 alu standoff kit", 1, "kit", "BUY", 18, "McMaster-Carr", "cad", "deck_gap 70 mm"),
    ("DR-AF-005", "Landing", "Landing gear carbon arch tube 22 mm + alu feet", "TaloNet GM-GEAR", 2, "set", "OUTSOURCE", 45, "Local CNC shop", "cad", "gear_height 230 mm"),
    ("DR-PR-001", "Propulsion", "Brushless motor, 14 kg-class octo, ~KV120", "T-Motor MN801S KV120", 8, "ea", "BUY", 175, "T-Motor", "docs/01", "X8 coaxial; 8 motors total"),
    ("DR-PR-002", "Propulsion", "Folding CF prop ~28-30 in (CW/CCW pair)", "T-Motor NS28x9.2 CF", 8, "ea", "BUY", 46, "T-Motor", "cad", "prop_dia ~711 mm"),
    ("DR-PR-003", "Propulsion", "ESC 80 A 12S HV, DShot600 + telemetry", "T-Motor FLAME 80A 12S", 8, "ea", "BUY", 95, "T-Motor", "docs/06", "1 per motor"),
]

DRONE_AVIONICS = [
    ("DR-AV-001", "Flight ctrl", "Flight controller, triple-IMU + carrier", "CubePilot Cube Orange+", 1, "ea", "BUY", 430, "CubePilot / IRLock", "docs/06", "ArduPilot; SERVO/RELAY outputs"),
    ("DR-AV-002", "GNSS", "GNSS receiver w/ Galileo OSNMA", "Septentrio mosaic-X5 board", 1, "ea", "BUY", 620, "Septentrio", "docs/06", "anti-spoof authenticated nav"),
    ("DR-AV-003", "GNSS", "Secondary GNSS/compass (CAN)", "Hex Here3+", 1, "ea", "BUY", 120, "CubePilot", "docs/06", "RAIM cross-check"),
    ("DR-AV-004", "Compute", "Companion compute module + carrier", "NVIDIA Jetson AGX Orin 32GB", 1, "ea", "BUY", 1199, "NVIDIA / Seeed", "docs/06", "EO/IR pipeline + defense/ ; VLM removed from loop"),
    ("DR-AV-005", "C2 radio", "Long-range telemetry radio pair, 900 MHz", "RFD900x (pair)", 1, "pr", "BUY", 200, "RFDesign", "docs/06", "MAVLink2 link to GCS"),
    ("DR-AV-006", "Camera", "EO/IR gimbal camera (FPV feed)", "Workswell/SIYI ZT6 EO+IR", 1, "ea", "BUY", 950, "SIYI", "docs/07", "RTSP/UDP to gcs cockpit"),
    ("DR-PW-001", "Power", "12S6P Li-ion / LiPo pack ~22 Ah", "Tattu 12S 22000 mAh", 1, "ea", "BUY", 520, "Tattu / GensAce", "docs/01", "endurance/range driver"),
    ("DR-PW-002", "Power", "Power module / current sensor 200 A", "Mauch HS-200-HV + PowerCube", 1, "ea", "BUY", 95, "Mauch", "docs/06", "I2C power monitor"),
    ("DR-PW-003", "Power", "5.3 V/8 A BEC (redundant, OR-ing)", "Mauch 5.3V/8A BEC", 2, "ea", "BUY", 28, "Mauch", "docs/06", "avionics bus"),
    ("DR-PW-004", "Power", "Main connectors + precharge (AS150/XT90-S)", "AS150 + XT90-S set", 1, "set", "BUY", 22, "Amass", "docs/06", "inrush limit"),
    ("DR-PW-005", "Power", "Fuse set: ANL 150 A + blade 3-10 A", "Littelfuse ANL/ATO kit", 1, "kit", "BUY", 24, "Littelfuse", "docs/06 §9", "per wire/fuse sizing"),
]

DRONE_PAYLOAD = [
    ("DR-NP-001", "Net aim", "Digital servo 20 kg (AIM-PAN, SERVO9)", "Savox SB-2290SG", 1, "ea", "BUY", 80, "Savox", "payload_map", "gimbal pan -60..+60 deg"),
    ("DR-NP-002", "Net aim", "Digital servo 20 kg (AIM-TILT, SERVO10)", "Savox SB-2290SG", 1, "ea", "BUY", 80, "Savox", "payload_map", "gimbal tilt 0..75 deg"),
    ("DR-NP-003", "Net aim", "Pan/tilt gimbal yoke + bearing (printed/CNC)", "TaloNet GM-GIMBAL", 1, "set", "MAKE", 35, "Garage 3D print", "cad", "net_launcher(); PETG-CF/alu"),
    ("DR-NP-004", "Net aim", "Servo UBEC 6 V / 5 A", "Castle 10A BEC Pro", 1, "ea", "BUY", 22, "Castle Creations", "docs/06", "servo bus, isolated"),
    ("DR-NP-005", "Cinch", "Brushed gearmotor 12 V (mouth cinch)", "Pololu 25D HP 12V", 1, "ea", "BUY", 40, "Pololu", "docs/06", "drawstring spool; SERVO11 driver"),
    ("DR-NP-006", "Cinch", "Motor driver H-bridge w/ current sense", "Pololu G2 24v13", 1, "ea", "BUY", 30, "Pololu", "docs/06", "cinch drive"),
    ("DR-NP-007", "Cinch", "Current sensor breakout", "INA226 module", 1, "ea", "BUY", 8, "Pololu/Adafruit", "docs/06", "tension/stall threshold"),
    ("DR-NP-008", "Cinch", "Drawstring spool + flanges (printed/turned)", "TaloNet GM-SPOOL", 1, "ea", "MAKE", 12, "Garage 3D print", "cad", "cinch_mechanism()"),
    ("DR-NP-009", "Winch", "Brushed gearmotor 12 V (hoist winch)", "Pololu 37D 12V", 1, "ea", "BUY", 40, "Pololu", "docs/06", "vertical recovery"),
    ("DR-NP-010", "Winch", "Motor driver H-bridge 21 A", "Pololu G2 24v21", 1, "ea", "BUY", 40, "Pololu", "docs/06", "winch drive"),
    ("DR-NP-011", "Winch", "Load cell 5 kg + HX711 ADC", "TAL220 + HX711", 1, "set", "BUY", 11, "SparkFun", "docs/06", "winch tension feedback"),
    ("DR-NP-012", "Trigger", "CO2 launcher: solenoid valve + 16 g cartridge", "12V NC CO2 solenoid", 1, "set", "BUY", 35, "Industrial supply", "docs/06", "RELAY0 fired"),
    ("DR-NP-013", "Trigger", "Logic-level MOSFET (low-side switch)", "Infineon IRLB3034PBF", 2, "ea", "BUY", 2, "Mouser", "docs/06", "solenoid/relay drive"),
    ("DR-NP-014", "Trigger", "Arming relay + flyback diode set", "Omron G5LE + 1N5408", 1, "set", "BUY", 9, "Mouser", "docs/06", "arming interlock"),
    ("DR-NP-015", "Release", "Quick-release servo (drop/abort)", "Savox SB-2290SG", 1, "ea", "BUY", 80, "Savox", "payload_map", "SERVO12 release"),
    ("DR-NP-016", "Net", "Capture net, knotless Dyneema/Kevlar, ~3.5 m mouth", "Custom cast net", 1, "ea", "MAKE", 60, "Garage / net maker", "cad", "capture_net(); launched casting net"),
    ("DR-NP-017", "Net", "Drawstring (Dyneema 2 mm) + perimeter weights", "Dyneema + brass weights", 1, "set", "MAKE", 18, "Garage", "cad", "purse loop + rim weights"),
    ("DR-NP-018", "Net", "Muzzle nozzle + bay housing (printed)", "TaloNet GM-MUZZLE", 1, "set", "MAKE", 20, "Garage 3D print", "cad", "netlauncher_bay()"),
    ("DR-NP-019", "Payload IF", "Payload interface PCB (STM32G474) fab + assembly", "TaloNet GM-PAYIF rev A", 1, "ea", "OUTSOURCE", 65, "JLCPCB/PCBWay", "docs/06 §10", "interlock + driver carrier"),
]

APPLIANCE = [
    ("FA-CP-001", "Compute", "Raspberry Pi Compute Module 4 (4GB/32GB)", "Raspberry Pi CM4104032", 1, "ea", "BUY", 75, "PiHut / DigiKey", "docs/09", "runs forensics/"),
    ("FA-CP-002", "Compute", "CM4 carrier board (USB3, NVMe, GbE, DSI)", "Waveshare CM4-IO-BASE-B", 1, "ea", "BUY", 45, "Waveshare", "docs/09", "I/O for appliance"),
    ("FA-UI-001", "Display", '7" capacitive touchscreen, 1024x600 DSI', "Waveshare 7inch DSI LCD", 1, "ea", "BUY", 70, "Waveshare", "docs/09", "operator UI"),
    ("FA-WB-001", "Write-block", "Hardware USB write-blocker (forensic)", "CRU WiebeTech USB WriteBlocker", 1, "ea", "BUY", 350, "CRU Inc.", "docs/09", "trust anchor; read-only"),
    ("FA-WB-002", "Card I/O", "USB SD/microSD reader (behind blocker)", "Kingston MobileLite Plus", 1, "ea", "BUY", 20, "Kingston", "docs/09", "card intake"),
    ("FA-PR-001", "Printer", "80 mm panel-mount ESC/POS thermal printer", "Rongta RP05 / CSN equiv", 1, "ea", "BUY", 60, "Rongta", "docs/09", "report output"),
    ("FA-PR-002", "Consumable", "80 mm thermal paper roll (pack of 4)", "Thermal 80x50 roll x4", 1, "pk", "BUY", 12, "Office supply", "docs/10", "report media"),
    ("FA-ST-001", "Storage", "NVMe SSD 512 GB (working/case store)", "Samsung 980 512GB", 1, "ea", "BUY", 50, "Samsung", "docs/09", "LUKS-encrypted"),
    ("FA-PW-001", "Power", "UPS HAT / DC-UPS + 18650 cells", "Geekworm X1200 + 2x 18650", 1, "set", "BUY", 45, "Geekworm", "docs/09", "clean shutdown / field run"),
    ("FA-PW-002", "Power", "AC/DC inlet + 12V PSU", "IEC inlet + 12V/5A PSU", 1, "set", "BUY", 22, "Mean Well", "docs/09", "mains"),
    ("FA-RT-001", "Time", "RTC module (UTC time source)", "DS3231 + coin cell", 1, "ea", "BUY", 6, "Adafruit", "docs/09", "trusted timestamps"),
    ("FA-EN-001", "Enclosure", "Rugged khaki enclosure (printed/CNC) + key-lock", "forensic_appliance.scad", 1, "ea", "MAKE", 80, "Garage 3D print", "cad", "ForensIQ-1 case; tamper lock"),
    ("FA-EN-002", "Controls", "Status LEDs, START/EJECT buttons, key-lock", "Panel control kit", 1, "kit", "BUY", 30, "Adafruit/AliExpress", "docs/10", "front panel"),
    ("FA-EX-001", "Export", "Write-once / WORM USB (evidence export)", "Kanguru WORM USB 32GB", 1, "ea", "BUY", 40, "Kanguru", "docs/10", "evidentiary PDF/map"),
    ("FA-CL-001", "Cooling", "60 mm fan + filtered intake", "Noctua NF-A6x25 5V", 1, "ea", "BUY", 16, "Noctua", "docs/09", "thermal"),
]

CONSUMABLES = [
    ("CO-WR-001", "Wire", "Silicone wire kit 10/12/14/18/20/22 AWG", "BNTECHGO silicone kit", 1, "kit", "BUY", 40, "BNTECHGO", "docs/06 §9", "per fuse/AWG sizing"),
    ("CO-WR-002", "Connectors", "XT60/XT90/JST-GH/Molex + crimp pins", "Amass + JST kit", 1, "kit", "BUY", 35, "Amass / JST", "docs/06", "looms"),
    ("CO-WR-003", "Insulation", "Heat shrink + braided sleeve + Kapton", "Assorted shrink kit", 1, "kit", "BUY", 18, "3M", "build", "harness"),
    ("CO-FS-001", "Fasteners", "M2/M2.5/M3 stainless socket-head + nylocs", "McMaster screw assortment", 1, "kit", "BUY", 45, "McMaster-Carr", "build", "airframe + payload"),
    ("CO-AD-001", "Adhesives", "Threadlocker, CA, 30-min epoxy, RTV", "Loctite 243 + JB Weld", 1, "kit", "BUY", 30, "Loctite", "build", "vibration-proofing"),
    ("CO-3D-001", "Filament", "PETG-CF / ASA filament (printed parts)", "Prusament PETG-CF 1kg", 2, "kg", "BUY", 40, "Prusa", "build", "gimbal/spool/muzzle/enclosure"),
    ("CO-TH-001", "Thermal", "Thermal pads/paste + standoffs (Jetson/CM4)", "Thermal kit", 1, "kit", "BUY", 15, "Arctic", "build", "compute cooling"),
]

OUTSOURCED = [
    ("OS-CN-001", "CNC", "CNC/waterjet carbon decks (2x) + alu mounts", "per cad STEP/STL", 1, "job", "OUTSOURCE", 0, "Local CNC / SendCutSend", "cad", "covered in DR-AF lines; quote per drawing"),
    ("OS-PC-001", "PCB", "Payload IF PCB fab + SMT assembly (STM32G474)", "TaloNet GM-PAYIF rev A", 1, "job", "OUTSOURCE", 0, "JLCPCB / PCBWay", "docs/06 §10", "covered in DR-NP-019; provide gerbers+BOM+CPL"),
    ("OS-AN-001", "Anodize", "Anodize/powder-coat alu parts (khaki #86855F)", "finish service", 1, "job", "OUTSOURCE", 60, "Local finisher", "cad", "signature khaki"),
]

TOOLS = [
    ("TL-001", "Solder", "Soldering station + hot air rework", "Hakko FX-951 / 8586", 1, "ea", "TOOL", 130, "Hakko", "shop", "harness + payload IF"),
    ("TL-002", "Crimp", "Ratchet crimpers (XT/JST/Molex/ferrule)", "IWISS crimp set", 1, "set", "TOOL", 70, "IWISS", "shop", "connectors"),
    ("TL-003", "3D print", "FDM 3D printer (enclosed, CF-capable)", "Bambu Lab X1C / Prusa MK4", 1, "ea", "TOOL", 1200, "Bambu Lab", "shop", "gimbal/spool/muzzle/enclosure (CNC alt)"),
    ("TL-004", "Measure", "Digital calipers + bench multimeter", "Mitutoyo + Fluke 117", 1, "set", "TOOL", 180, "Mitutoyo/Fluke", "shop", "QA"),
    ("TL-005", "Power", "Bench PSU 30V/10A + LiPo charger", "Riden + ISDT", 1, "set", "TOOL", 150, "Riden / ISDT", "shop", "bring-up + battery"),
    ("TL-006", "Hand", "Torque drivers, hex set, helping hands, heat gun", "Wera + misc", 1, "set", "TOOL", 120, "Wera", "shop", "assembly"),
    ("TL-007", "Safety", "LiPo-safe bag/bunker, CO2-safe handling, fume ext.", "Safety kit", 1, "set", "TOOL", 90, "misc", "shop", "battery + CO2 + solder fumes"),
]

SECTIONS = [
    ("Drone — Airframe & Propulsion", DRONE_FRAME),
    ("Drone — Avionics, Power & C2", DRONE_AVIONICS),
    ("Drone — Net Payload", DRONE_PAYLOAD),
    ("Forensic Appliance (ForensIQ-1)", APPLIANCE),
    ("Fasteners, Wiring & Consumables", CONSUMABLES),
    ("Outsourced Fabrication", OUTSOURCED),
    ("Shop Tools & Equipment", TOOLS),
]


def style_header(ws):
    for c, (name, width) in enumerate(zip(COLS, WIDTHS), start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=KHAKI)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def write_sheet(wb, title, rows):
    ws = wb.create_sheet(title[:31])
    style_header(ws)
    r = 2
    for (pid, cat, item, mpn, qty, unit, disp, price, vendor, ref, notes) in rows:
        vals = [pid, cat, item, mpn, qty, unit, disp, price, None, vendor, ref, notes]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center",
                                       wrap_text=(c in (3, 12)),
                                       horizontal="center" if c in (5, 6, 7) else "left")
        ws.cell(row=r, column=8).number_format = '$#,##0'
        ext = ws.cell(row=r, column=9, value=f"=E{r}*H{r}")
        ext.number_format = '$#,##0'
        ext.border = BORDER
        dc = ws.cell(row=r, column=7)
        dc.fill = PatternFill("solid", fgColor=DISP_FILL.get(disp, "FFFFFF"))
        dc.font = Font(bold=True, size=10)
        r += 1
    total_row = r
    ws.cell(row=total_row, column=8, value="SUBTOTAL").font = Font(bold=True)
    ws.cell(row=total_row, column=8).alignment = Alignment(horizontal="right")
    tc = ws.cell(row=total_row, column=9, value=f"=SUM(I2:I{total_row - 1})")
    tc.font = Font(bold=True)
    tc.number_format = '$#,##0'
    tc.fill = PatternFill("solid", fgColor=KHAKI)
    tc.font = Font(bold=True, color="FFFFFF")
    return f"'{ws.title}'!I{total_row}"


def cover_sheet(wb):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 90
    def line(r, text, size=11, bold=False, color="222222"):
        c = ws.cell(row=r, column=2, value=text)
        c.font = Font(size=size, bold=bold, color=color)
        c.alignment = Alignment(wrap_text=True, vertical="top")
    line(2, "TaloNet — Bill of Materials & Procurement Package", 18, True, KHAKI)
    line(3, "그물매 net-interceptor drone + ForensIQ-1 forensic appliance")
    line(5, "HOW TO USE", 12, True)
    line(6, "One sheet per subsystem. Disp = MAKE (garage) / BUY (off-the-shelf) / "
            "OUTSOURCE (CNC/PCB/finish) / TOOL (one-time shop). Ext USD = Qty x Unit "
            "(formula). 'Cost Roll-up' sums every sheet. Prices are INDICATIVE USD — "
            "confirm with the vendor before ordering.")
    line(8, "LEGEND", 12, True)
    for i, (k, txt) in enumerate([
            ("MAKE", "fabricate/assemble in the garage (3D print, net rigging, wiring)"),
            ("BUY", "purchase off-the-shelf"),
            ("OUTSOURCE", "send a drawing/gerber out (CNC carbon, PCB fab+assembly, anodize)"),
            ("TOOL", "one-time shop equipment, not part of a single airframe cost")]):
        cell = ws.cell(row=9 + i, column=2, value=f"   {k}  —  {txt}")
        cell.font = Font(size=11)
        ws.cell(row=9 + i, column=1).fill = PatternFill(
            "solid", fgColor=DISP_FILL[k])
    line(14, "TRACE", 12, True)
    line(15, "Part numbers trace to docs/01 (drone spec), docs/06 (circuit BOM + "
             "pinout), docs/09 (appliance), and gcs/payload_map.py (servo channels). "
             "Build/assembly: docs/12_Garage_Build_Guide.md.")
    line(17, "SAFETY / LEGAL", 12, True, color="B02828")
    line(18, "High-energy LiPo, pressurized CO2, and a kinetic net launcher are no "
             "joke — see the safety section of docs/12. Counter-UAS use is for "
             "authorized operators only; the system is intelligence/control + a "
             "human-in-the-loop trigger, never an autonomous weapon.")
    line(20, "Generated by tools/generate_bom.py — edit + re-run to update.", 9, False, "888888")


def rollup_sheet(wb, totals):
    ws = wb.create_sheet("Cost Roll-up")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 16
    hdr = ["Subsystem", "Subtotal USD"]
    for c, name in enumerate(hdr, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=KHAKI)
        cell.border = BORDER
    r = 2
    first = r
    for title, ref in totals:
        ws.cell(row=r, column=1, value=title).border = BORDER
        v = ws.cell(row=r, column=2, value=f"={ref}")
        v.number_format = '$#,##0'
        v.border = BORDER
        r += 1
    last = r - 1
    ws.cell(row=r, column=1, value="Materials subtotal (incl. tools)").font = Font(bold=True)
    sub = ws.cell(row=r, column=2, value=f"=SUM(B{first}:B{last})")
    sub.number_format = '$#,##0'
    sub.font = Font(bold=True)
    ws.cell(row=r + 1, column=1, value="Contingency 15%")
    cont = ws.cell(row=r + 1, column=2, value=f"=B{r}*0.15")
    cont.number_format = '$#,##0'
    ws.cell(row=r + 2, column=1, value="GRAND TOTAL (indicative)").font = Font(bold=True, size=12)
    gt = ws.cell(row=r + 2, column=2, value=f"=B{r}+B{r + 1}")
    gt.number_format = '$#,##0'
    gt.font = Font(bold=True, color="FFFFFF", size=12)
    gt.fill = PatternFill("solid", fgColor=KHAKI)


def main():
    wb = Workbook()
    wb.remove(wb.active)
    totals = []
    for title, rows in SECTIONS:
        ref = write_sheet(wb, title, rows)
        totals.append((title, ref))
    cover_sheet(wb)
    rollup_sheet(wb, totals)
    out_dir = os.path.join(os.path.dirname(__file__), "..", "bom")
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.normpath(os.path.join(out_dir, "TaloNet_BOM.xlsx"))
    wb.save(out)
    n = sum(len(rows) for _, rows in SECTIONS)
    print(f"wrote {out}  ({len(SECTIONS)} sheets, {n} line items)")


if __name__ == "__main__":
    main()
